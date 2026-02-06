import streamlit as st
import openpyxl
import pandas as pd
from io import BytesIO

# Page configuration
st.set_page_config(
    page_title="Monthly Review Requests",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for glassy premium design with emerald and gold
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Inter:wght@300;400;500;600&display=swap');
    
    /* Main app styling */
    .stApp {
        background: linear-gradient(135deg, #064e3b 0%, #065f46 25%, #047857 50%, #10b981 75%, #34d399 100%);
        font-family: 'Inter', sans-serif;
        animation: gradientShift 15s ease infinite;
        background-size: 200% 200%;
    }
    
    @keyframes gradientShift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Hide Fork button and toolbar menu */
    [data-testid="stToolbar"] {visibility: hidden;}
    .stDeployButton {visibility: hidden;}
    [data-testid="stDecoration"] {visibility: hidden;}
    button[kind="header"] {visibility: hidden;}
    
    /* Main container */
    .main .block-container {
        padding: 2rem 3rem;
        max-width: 1400px;
        animation: fadeIn 0.8s ease-in;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    /* Glass card effect */
    .glass-card {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        padding: 2rem;
        margin: 1rem 0;
        box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.37);
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    .glass-card:hover {
        transform: translateY(-5px) scale(1.02);
        box-shadow: 0 15px 50px rgba(16, 185, 129, 0.4);
        border: 1px solid rgba(236, 201, 75, 0.4);
    }
    
    /* Header styling */
    .app-header {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(236, 201, 75, 0.2));
        backdrop-filter: blur(15px);
        border-radius: 25px;
        border: 2px solid rgba(236, 201, 75, 0.3);
        padding: 2.5rem;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 10px 40px rgba(236, 201, 75, 0.2);
        animation: headerFloat 3s ease-in-out infinite;
        transition: all 0.4s ease;
    }
    
    .app-header:hover {
        transform: scale(1.03);
        box-shadow: 0 15px 60px rgba(236, 201, 75, 0.4);
    }
    
    @keyframes headerFloat {
        0%, 100% { transform: translateY(0px); }
        50% { transform: translateY(-10px); }
    }
    
    .app-title {
        font-family: 'Playfair Display', serif;
        font-size: 3.5rem;
        font-weight: 700;
        background: linear-gradient(135deg, #ecc94b 0%, #f6e05e 50%, #faf089 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 0.5rem;
        text-shadow: 0 0 30px rgba(236, 201, 75, 0.3);
        animation: titlePulse 2s ease-in-out infinite;
    }
    
    @keyframes titlePulse {
        0%, 100% { filter: brightness(1); }
        50% { filter: brightness(1.2); }
    }
    
    .app-subtitle {
        font-family: 'Inter', sans-serif;
        font-size: 1.2rem;
        color: #d1fae5;
        font-weight: 300;
        letter-spacing: 2px;
        animation: subtitleSlide 1s ease-out;
    }
    
    @keyframes subtitleSlide {
        from { opacity: 0; transform: translateX(-30px); }
        to { opacity: 1; transform: translateX(0); }
    }
    
    /* Request card styling */
    .request-card {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.15), rgba(255, 255, 255, 0.05));
        backdrop-filter: blur(12px);
        border-radius: 18px;
        border: 1px solid rgba(236, 201, 75, 0.2);
        padding: 2rem;
        margin: 1.5rem 0;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.2);
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        animation: slideIn 0.6s ease-out;
    }
    
    @keyframes slideIn {
        from { 
            opacity: 0; 
            transform: translateX(-50px);
        }
        to { 
            opacity: 1; 
            transform: translateX(0);
        }
    }
    
    .request-card:hover {
        transform: translateY(-8px) scale(1.02);
        box-shadow: 0 20px 60px rgba(236, 201, 75, 0.5);
        border: 1px solid rgba(236, 201, 75, 0.6);
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.2), rgba(255, 255, 255, 0.08));
    }
    
    .request-number {
        font-family: 'Playfair Display', serif;
        font-size: 1.8rem;
        font-weight: 700;
        color: #fbbf24;
        margin-bottom: 1rem;
        text-shadow: 0 0 10px rgba(251, 191, 36, 0.5);
        transition: all 0.3s ease;
        display: inline-block;
    }
    
    .request-card:hover .request-number {
        transform: scale(1.15);
        text-shadow: 0 0 30px rgba(251, 191, 36, 1), 0 0 40px rgba(251, 191, 36, 0.8), 0 0 50px rgba(251, 191, 36, 0.6);
        filter: brightness(1.3);
    }
    
    /* Upload section */
    .upload-section {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.15), rgba(52, 211, 153, 0.1));
        backdrop-filter: blur(10px);
        border-radius: 20px;
        border: 2px dashed rgba(236, 201, 75, 0.4);
        padding: 3rem;
        text-align: center;
        margin: 2rem 0;
        transition: all 0.4s ease;
        animation: pulse 3s ease-in-out infinite;
    }
    
    @keyframes pulse {
        0%, 100% { border-color: rgba(236, 201, 75, 0.4); }
        50% { border-color: rgba(236, 201, 75, 0.7); }
    }
    
    .upload-section:hover {
        transform: scale(1.03);
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.25), rgba(52, 211, 153, 0.15));
        border: 2px dashed rgba(236, 201, 75, 0.7);
        box-shadow: 0 10px 40px rgba(16, 185, 129, 0.3);
    }
    
    /* Buttons */
    .stDownloadButton button {
        background: linear-gradient(135deg, #10b981, #34d399);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    
    .stDownloadButton button::before {
        content: '';
        position: absolute;
        top: 50%;
        left: 50%;
        width: 0;
        height: 0;
        border-radius: 50%;
        background: rgba(255, 255, 255, 0.3);
        transform: translate(-50%, -50%);
        transition: width 0.6s, height 0.6s;
    }
    
    .stDownloadButton button:hover::before {
        width: 300px;
        height: 300px;
    }
    
    .stDownloadButton button:hover {
        background: linear-gradient(135deg, #059669, #10b981);
        box-shadow: 0 8px 30px rgba(16, 185, 129, 0.8);
        transform: translateY(-3px) scale(1.05);
    }
    
    .stDownloadButton button:active {
        transform: translateY(-1px) scale(1.02);
    }
    
    /* Divider */
    .gold-divider {
        height: 2px;
        background: linear-gradient(90deg, transparent, #fbbf24, transparent);
        margin: 2rem 0;
        box-shadow: 0 0 10px rgba(251, 191, 36, 0.3);
        animation: dividerGlow 2s ease-in-out infinite;
    }
    
    @keyframes dividerGlow {
        0%, 100% { opacity: 0.7; }
        50% { opacity: 1; box-shadow: 0 0 20px rgba(251, 191, 36, 0.6); }
    }
    
    /* Stats cards */
    .stat-card {
        background: linear-gradient(135deg, rgba(251, 191, 36, 0.2), rgba(236, 201, 75, 0.1));
        backdrop-filter: blur(10px);
        border-radius: 15px;
        border: 1px solid rgba(251, 191, 36, 0.3);
        padding: 1.5rem;
        text-align: center;
        margin: 1rem 0;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        animation: statFloat 3s ease-in-out infinite;
    }
    
    .stat-card:hover {
        transform: translateY(-10px) rotate(2deg) scale(1.08);
        box-shadow: 0 15px 40px rgba(251, 191, 36, 0.5);
        border: 1px solid rgba(251, 191, 36, 0.7);
        background: linear-gradient(135deg, rgba(251, 191, 36, 0.3), rgba(236, 201, 75, 0.15));
    }
    
    @keyframes statFloat {
        0%, 100% { transform: translateY(0); }
        50% { transform: translateY(-5px); }
    }
    
    .stat-number {
        font-family: 'Playfair Display', serif;
        font-size: 2.5rem;
        font-weight: 700;
        color: #fbbf24;
        text-shadow: 0 0 20px rgba(251, 191, 36, 0.5);
        transition: all 0.3s ease;
    }
    
    .stat-card:hover .stat-number {
        transform: scale(1.2);
        text-shadow: 0 0 30px rgba(251, 191, 36, 0.9);
    }
    
    .stat-label {
        color: #d1fae5;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 2px;
        margin-top: 0.5rem;
        transition: all 0.3s ease;
    }
    
    .stat-card:hover .stat-label {
        letter-spacing: 3px;
        color: #ffffff;
    }
    
    /* File uploader */
    .stFileUploader {
        background: rgba(255, 255, 255, 0.05);
        border-radius: 15px;
        padding: 1rem;
        transition: all 0.3s ease;
    }
    
    .stFileUploader:hover {
        background: rgba(255, 255, 255, 0.1);
        transform: scale(1.02);
    }
    
    /* Success/Error messages */
    .stSuccess, .stError, .stWarning {
        background: rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        border-radius: 12px;
        animation: messageSlide 0.5s ease-out;
    }
    
    @keyframes messageSlide {
        from { 
            opacity: 0; 
            transform: translateX(-20px);
        }
        to { 
            opacity: 1; 
            transform: translateX(0);
        }
    }
    
    /* Code block styling (for copy button) */
    .stCodeBlock {
        transition: all 0.3s ease;
        border-radius: 8px;
        overflow: hidden;
    }
    
    .stCodeBlock:hover {
        transform: scale(1.01);
        box-shadow: 0 5px 20px rgba(16, 185, 129, 0.3);
    }
    
    /* Markdown content animations */
    .stMarkdown {
        animation: contentFadeIn 0.6s ease-out;
    }
    
    @keyframes contentFadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    /* Scrollbar styling */
    ::-webkit-scrollbar {
        width: 12px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(0, 0, 0, 0.2);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #10b981, #34d399);
        border-radius: 10px;
        transition: all 0.3s ease;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #059669, #10b981);
        box-shadow: 0 0 10px rgba(16, 185, 129, 0.5);
    }
</style>
""", unsafe_allow_html=True)

def extract_and_format_requests(uploaded_file):
    """
    Read Excel file and return formatted monthly review requests
    """
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active

        # Get the header row (row 5)
        headers = []
        for cell in ws[5]:
            headers.append(cell.value)

        # Column indices (0-based)
        name_idx = 0  # A
        year_idx = 5  # F
        subject_idx = 7  # H
        requested_time_idx = 13  # N
        available_times_idx = 14  # O
        notes_idx = 16  # Q

        # Process each data row
        data_rows = list(ws.iter_rows(min_row=6, values_only=True))
        requests = []

        for idx, row in enumerate(data_rows, 1):
            if row[0] is None:  # Skip empty rows
                continue

            student_name = row[name_idx] if name_idx < len(row) else ""
            year = row[year_idx] if year_idx < len(row) else ""
            subject = row[subject_idx] if subject_idx < len(row) else ""
            requested_time = row[requested_time_idx] if requested_time_idx < len(row) else ""
            available_times = row[available_times_idx] if available_times_idx < len(row) else ""
            notes = row[notes_idx] if notes_idx < len(row) else ""

            # Skip rows with missing critical data
            if not student_name or not year or not subject:
                continue

            # Use available times if requested time is not specified
            time_slot = requested_time if requested_time else available_times

            requests.append({
                'number': len(requests) + 1,
                'student_name': student_name,
                'year': year,
                'subject': subject,
                'time_slot': time_slot,
                'notes': notes if notes else "No specific notes"
            })

        return requests, None

    except Exception as e:
        return None, str(e)

def display_request_card(request):
    """Display a single request in text format with copy button"""
    st.markdown(f"""
    <div class="request-card">
        <div class="request-number"><strong>📋 Request #{request['number']} - Year {request['year']} {request['subject']}</strong></div>
    </div>
    """, unsafe_allow_html=True)
    
    # Create the message text
    message_text = f"""We'd love to allocate a Monthly Review lesson to you!
Here are the details:

👤 Student: {request['student_name']}
📚 Year and Subject: Year {request['year']} - {request['subject']}
🕒 Preferred Time Slot: {request['time_slot']}
📝 Topics or Notes: {request['notes']}

Please let us know if you're available to take this class.
✅ If you're happy with the time slot, we'll proceed with the setup.
⏳ If not, kindly suggest an alternative time and we'll confirm with the parent.

Looking forward to your response!"""
    
    # Display in black box for easy reading
    # st.markdown(f"""
    # <div style="background-color: #1a1a1a; padding: 1.5rem; border-radius: 12px; margin: 1rem 0; color: #ffffff; font-family: 'Inter', sans-serif; line-height: 1.8;">
    #     <pre style="margin: 0; white-space: pre-wrap; font-family: 'Inter', sans-serif; color: #ffffff;">{message_text}</pre>
    # </div>
    # """, unsafe_allow_html=True)
    
    # Copy button
    st.code(message_text, language=None)
    
    st.markdown("---")

def create_text_export(requests):
    """Create formatted text for export"""
    text_output = ""
    for request in requests:
        text_output += f"""We'd love to allocate a Monthly Review lesson to you! 
Here are the details:
👤 Student: {request['student_name']}
📚 Year and Subject: Year {request['year']} - {request['subject']}
🕒 Preferred Time Slot: {request['time_slot']}
📝 Topics or Notes: {request['notes']}
Please let us know if you're available to take this class.
✅ If you're happy with the time slot, we'll proceed with the setup.
⏳ If not, kindly suggest an alternative time and we'll confirm with the parent.
Looking forward to your response!

{"=" * 80}

"""
    return text_output

# Main app
def main():
    # Header
    st.markdown("""
    <div class="app-header">
        <div class="app-title">📚 Monthly Review Requests Portal</div>
        <div class="app-subtitle">Premium Request Management System By Mohammed Abdelwahed</div>
    </div>
    """, unsafe_allow_html=True)

    # Upload section
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    st.markdown("### ✨ Upload Your Excel File")
    st.markdown("Upload your monthly review requests Excel file to get started")
    uploaded_file = st.file_uploader("", type=['xlsx', 'xls'], label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_file is not None:
        with st.spinner('🔄 Processing your requests...'):
            requests, error = extract_and_format_requests(uploaded_file)

        if error:
            st.error(f"❌ Error processing file: {error}")
        elif requests:
            # Display stats
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{len(requests)}</div>
                    <div class="stat-label">Total Requests</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                subjects = set([r['subject'] for r in requests if r['subject']])
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{len(subjects)}</div>
                    <div class="stat-label">Unique Subjects</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                years = set([r['year'] for r in requests if r['year']])
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{len(years)}</div>
                    <div class="stat-label">Year Groups</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)

            # Export button
            text_export = create_text_export(requests)
            st.download_button(
                label="📥 Download All Requests as Text",
                data=text_export,
                file_name="monthly_review_requests.txt",
                mime="text/plain"
            )

            st.markdown('<div class="gold-divider"></div>', unsafe_allow_html=True)

            # Display requests
            st.markdown("### 📋 All Requests")
            for request in requests:
                display_request_card(request)
        else:
            st.warning("⚠️ No requests found in the uploaded file.")
    else:
        # Welcome message
        st.markdown("""
        <div class="glass-card" style="text-align: center; padding: 3rem;">
            <h2 style="color: #fbbf24; font-family: 'Playfair Display', serif; margin-bottom: 1rem;">
                Welcome to Monthly Review Requests Manager
            </h2>
            <p style="color: #d1fae5; font-size: 1.1rem; line-height: 1.8;">
                Upload your Excel file to view beautifully formatted monthly review requests.<br>
                The system will automatically extract and display all student information,<br>
                including preferred time slots and special notes.
            </p>
        </div>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
